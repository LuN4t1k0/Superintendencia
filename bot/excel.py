from io import BytesIO

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def read_ruts(file_bytes: bytes) -> tuple[Workbook, Worksheet, int, list[int]]:
    """Load workbook and return workbook, active sheet, RUT column, and data rows."""
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active

    rut_col = None
    for cell in next(ws.iter_rows(1, 1)):
        if cell.value and "rut" in str(cell.value).lower():
            rut_col = cell.column
            break

    if rut_col is None:
        raise ValueError("No se encontró columna con 'RUT' en el header")

    data_rows = [
        row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, rut_col).value is not None
    ]
    return wb, ws, rut_col, data_rows


def find_or_create_afp2_column(ws: Worksheet) -> int:
    """Return the 1-based AFP 2 column index, creating it when absent."""
    for cell in next(ws.iter_rows(1, 1)):
        if cell.value and str(cell.value).strip().upper() == "AFP 2":
            return cell.column

    new_col = ws.max_column + 1
    ws.cell(1, new_col, "AFP 2")
    return new_col


def get_pending_rows(
    ws: Worksheet, rut_col: int, afp2_col: int, data_rows: list[int]
) -> list[int]:
    """Return rows where AFP 2 is still empty (not yet processed)."""
    return [row for row in data_rows if not ws.cell(row, afp2_col).value]


def write_afp(ws: Worksheet, row: int, col: int, value: str) -> None:
    ws.cell(row, col, value)


def to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
