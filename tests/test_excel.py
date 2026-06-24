import io
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from bot.excel import find_or_create_afp2_column, get_pending_rows, read_ruts, to_bytes, write_afp


def _make_excel(headers: list, rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_read_ruts_detecta_columna_rut():
    data = _make_excel(
        ["Nombre", "RUT", "AFP"],
        [["Juan", "15800185-3", ""], ["Maria", "12345678-9", ""]],
    )
    wb, ws, rut_col, data_rows = read_ruts(data)
    assert wb is not None
    assert ws is not None
    assert rut_col == 2
    assert data_rows == [2, 3]


def test_read_ruts_case_insensitive():
    data = _make_excel(["Nombre", "Rut Trabajador"], [["Juan", "15800185-3"]])
    wb, ws, rut_col, data_rows = read_ruts(data)
    assert wb is not None
    assert ws is not None
    assert rut_col == 2
    assert data_rows == [2]


def test_read_ruts_sin_columna_rut_lanza_error():
    data = _make_excel(["Nombre", "AFP"], [["Juan", "HABITAT"]])
    with pytest.raises(ValueError, match="No se encontró"):
        read_ruts(data)


def test_read_ruts_ignora_filas_vacias():
    data = _make_excel(
        ["RUT"],
        [["15800185-3"], [None], ["12345678-9"]],
    )
    wb, ws, rut_col, data_rows = read_ruts(data)
    assert wb is not None
    assert ws is not None
    assert rut_col == 1
    assert data_rows == [2, 4]


def test_find_or_create_afp2_existente():
    wb = Workbook()
    ws = wb.active
    ws.append(["RUT", "AFP 2", "Nombre"])
    assert find_or_create_afp2_column(ws) == 2


def test_find_or_create_afp2_crea_nueva():
    wb = Workbook()
    ws = wb.active
    ws.append(["RUT", "Nombre"])
    col = find_or_create_afp2_column(ws)
    assert col == 3
    assert ws.cell(1, 3).value == "AFP 2"


def test_write_afp():
    wb = Workbook()
    ws = wb.active
    ws.append(["RUT", "AFP 2"])
    ws.append(["15800185-3", ""])
    write_afp(ws, 2, 2, "HABITAT")
    assert ws.cell(2, 2).value == "HABITAT"


def test_get_pending_rows_filtra_ya_procesados():
    wb = Workbook()
    ws = wb.active
    ws.append(["RUT", "AFP 2"])
    ws.append(["15800185-3", "HABITAT"])   # ya procesado
    ws.append(["12345678-9", ""])           # pendiente
    ws.append(["11111111-1", None])         # pendiente
    pending = get_pending_rows(ws, rut_col=1, afp2_col=2, data_rows=[2, 3, 4])
    assert pending == [3, 4]


def test_get_pending_rows_todos_pendientes():
    wb = Workbook()
    ws = wb.active
    ws.append(["RUT", "AFP 2"])
    ws.append(["15800185-3", None])
    ws.append(["12345678-9", None])
    pending = get_pending_rows(ws, rut_col=1, afp2_col=2, data_rows=[2, 3])
    assert pending == [2, 3]


def test_get_pending_rows_todos_procesados():
    wb = Workbook()
    ws = wb.active
    ws.append(["RUT", "AFP 2"])
    ws.append(["15800185-3", "HABITAT"])
    ws.append(["12345678-9", "CAPITAL"])
    pending = get_pending_rows(ws, rut_col=1, afp2_col=2, data_rows=[2, 3])
    assert pending == []


def test_to_bytes_roundtrip():
    wb = Workbook()
    wb.active.append(["RUT", "AFP 2"])
    wb.active.append(["15800185-3", "HABITAT"])
    raw = to_bytes(wb)
    wb2 = load_workbook(io.BytesIO(raw))
    assert wb2.active.cell(2, 2).value == "HABITAT"
